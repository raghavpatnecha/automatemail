import os.path
import openpyxl
from string import Template
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.mime.base import MIMEBase
from optparse import OptionParser
from email.mime.image import MIMEImage
from collections import defaultdict

excel_file = os.path.join('#path to file', 'example.xlsx')           # goal is to get excel file
msg_template = os.path.join('#path to file', 'invite.txt')           # goal is to get message template file
id_file = os.path.join('#path to file', 'id.txt')                    # goal is to get id file
add_attachment = '#path to file'                              #add attachment here e.g. pdf, excel, word, image
parser = OptionParser()                                             #initialize parser to get input from cmd or bash
parser.add_option("-f", "--from", dest="sender", help="sender email address", default="raghavpatnecha@outlook.com")     #add email as shown in the example image otherwise use the default one
parser.add_option("-i", "--image", dest="image", help="image attachment", default=False)                #add image signature (optional)
(options, args) = parser.parse_args()                                                                   #excecuting parser variables
email_username = options.sender                                                                       # goal is to get  email from the parser
email_password = 'add password'                                                             # add your email password if needed otherwise comment it


def get_contacts(filename):                                                      #get_contact function responsible for getting coumn names and respective values from xlsx file
    wb = openpyxl.load_workbook(filename)                                         #getting excel file name from variable
    sheet = wb['Sheet1']                                                          #adding excel sheet name
    try:
        id = open(id_file).read().splitlines()    #goal is to get id from txt file
        id = list(map(int, id))                   #inserting id in list
    except Exception or ImportError:
        senderrormail()                            #if file not found or error then send error mail
    values = []                                    #creating empty list
    dict = {}                                       #creating empty dictionary
    header = []                                     #creating empty list
    for i in range(2, sheet.max_row + 1):
        header.append(sheet.cell(row=1, column=i).value)        #goal is to insert excel header names i.e column names to the empty list header

    for rowidx in range(2, sheet.max_row + 1):                     #this loop stores
        row = int(sheet['A' + str(rowidx)].value)                  #get all the values in the id column
        if row in id:                                              #filter on the basis of id in out id.txt file
            show = sheet[str(rowidx)]                              #get rows object e.g  <Cell u'Sheet1'.A2>
            key = sheet.cell(row=rowidx, column=1).value           #storing filtered id in the variable key
            #key1 = sheet.cell(row=rowidx, column=2).value
            for cells in show:                                         #goal for this loop is to obtain values from row object and sore them in the dictionary
                values.append(cells.value)
                dict[key] = values[1:]                                  #inserting in dictionary
            del values[:]                                               #deleting first instance of dictionary in order to store a new key,value pair i.e {id:[Jhon, jhon@gmail.com, smith]}

    for i in range(len(header)):                                #this loop basically helps in mapping column names to the dictionary values
        d = {k: {name: v[i] for i, name in enumerate(header)} for k, v in dict.items()}   # so basically now the dictionary looks like this e.g {10:{'Owner Name': 'abc','Email': 'wa@gmail.com', 'Manager Name': 'jhon'} which makes it easy to access

    results = defaultdict(list)
    for j in id:    # this loop is responsible for storing values in defaultdict
        obj = d[j]
        for key, value in obj.items(): # the goal is to club all the owner name , manager name, email etc. together
            results[key].append(value)   #the output looks like this defaultdict(<type 'list'>, {u'Owner Name': [u'abc', u'awdaw'], u'Manager Name': [u'jhon', u'raghav']})
    return results                       #returning the results i.e default dict to the main function


def read_template(filename):                              # read_template function is responsible for reading the .oft file
    filepath = os.path.join(filename)
    with open(filepath, 'r') as template_file:             #opening the file
        template_file_content = template_file.read()        #reading contents of the file
    return Template(template_file_content)                  #  returning it to the main function


def senderrormail():                                    # the objective for this function is to send an email if an error is occured during the program execution
    try:
        smtpObj = smtplib.SMTP('smtp-mail.outlook.com', 587)  #imap.gmail.com for gmail
        smtpObj.starttls()
        sender = email_username
        sender_password = "add password"                   #type password if any otherwise comment this line
        receivers = 'raghavpatnecha06@gmail.com'            #receivers email address
        smtpObj.login(sender, sender_password)
        msg = MIMEMultipart()
        message = 'File not found'                      #type your message here
        msg['From'] = sender
        msg['To'] = receivers
        msg['Subject'] = "Excel file not found"                #change subject of the message
        msg.attach(MIMEText(message, 'plain'))
        #smtpObj.send_message(msg)
        smtpObj.sendmail(sender, receivers, msg.as_string())
        del msg
        print("Error email sent successfully")
    except smtplib.SMTPException:
        print("Error: unable to send email")

def main():                                                 # this is the first function which executes when our script runs

    try:
        message_template = read_template(msg_template)                  # passing the file name stored in msg_template variable to the read_template function
    except Exception or ImportError:
        senderrormail()                                                 #if file not found or import error then send an email

    # set up the SMTP server
    s = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)  # if sending mail through  smtp-mail.outlook.com

    try:
        #s.set_debuglevel(1)
        s.starttls()                                        # establishing the connection with the server
    except smtplib.SMTPException:
        pass
    s.login(email_username, email_password)                   #logging to mail server using supplied email

    try:
        results = get_contacts(excel_file)  # read contacts from the excel file stored in excel_file variable
        names, emails, managers = results["Owner Name"], results["Email"],results['Manager Name']  #add or remove column if you want #position 1 the variables results["Owner Name"],results["Email"] are the column names in the excel file

    except Exception or ImportError:      #send email if file not found
        senderrormail()
    while True:
        for name, email, manager in zip(names, emails, managers):   ## For each contact, send the email: #position 2  zipping all the names, emails, manager togerther to iterate over one by one

            msg = MIMEMultipart()  # create a message
            # add in the actual person name to the message template
            message = message_template.substitute(PERSON_NAME=name.title(),MANAGER_NAME=manager.title())   #position 3  #include what you want to include in template, the variable PERSON_NAME, MANAGER_NAME are the name stored in our template .oft file and the manger.title(), name.title() are the variables which we got from the zipping of names, emails, managers
            # Prints out the message body for our sake
            print(message)
            # setup the parameters of the message
            msg['From'] = email_username
            msg['To'] = email
            msg['Subject'] = "This is an test email"  # Subject of Email
            msg.attach(MIMEText(message, 'html'))

            if options.image is not False:         #goal is to add image(signature) as shown in the example image i.e example_how_to_send.png
                img = open(options.image, 'rb').read()
                msgImg = MIMEImage(img, 'png')    #reading the image as MIMEImage format
                msgImg.add_header('Content-ID', '<image1>')            #adding image data to the header
                msgImg.add_header('Content-Disposition', 'inline', filename=options.image)      # specifying to attach image as inline i.e inside the body of the mail
                msg.attach(msgImg)                                                                  #attaching image to the message

            try:
                f = add_attachment     #add attachment
                part = MIMEBase('application', "octet-stream")
                part.set_payload(open(f, "rb").read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(f))   #specifying the attachment in the header of the mail
                msg.attach(part)          #adding attachment to the message
            except Exception or ImportError:
                senderrormail()

            # send the message via the server set up earlier.
            #s.send_message(msg)
            s.sendmail(email_username, email, msg.as_string())        #finally sending email

            del msg

        # Terminate the SMTP session and close the connection
        s.close()
        s.quit()


if __name__ == '__main__':

    main()                                   # telling the script to execute the main function
